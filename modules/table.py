from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Inline logo (base64-encoded, offline-safe) ───────────────────────────────
_LOGO_B64_54 = "iVBORw0KGgoAAAANSUhEUgAAADYAAAA2CAYAAACMRWrdAAAAAXNSR0IArs4c6QAAAHhlWElmTU0AKgAAAAgABAEaAAUAAAABAAAAPgEbAAUAAAABAAAARgEoAAMAAAABAAIAAIdpAAQAAAABAAAATgAAAAAAAABIAAAAAQAAAEgAAAABAAOgAQADAAAAAQABAACgAgAEAAAAAQAAADagAwAEAAAAAQAAADYAAAAAwMwakwAAAAlwSFlzAAALEwAACxMBAJqcGAAAFtFJREFUaAXtWntwXNV9/u7d90paPVdPy/JDlizb+CVjg8E2GPM0BJMBkqbNiyZOJ5k0UzJN0ukkQ/9oZzJJSWbSdCbTPAZKSIMDpMFAgBhsSIxf4Jf8lC15bb12tVqt9r17d+/2+527a1mSoe1f7R8c+ey999zz+H2/9znXwEflIw78v+CA9r+hovjcc7bdB0/Xvvve8eapolGVshU1Awby+Txg5JEzclo+nZcWIC/tMntenlQpFotqPXnOa1pRGh12u3rHO9g9dtj5rDmcRYfHDU3Tix5dTy9Z2BX8x5/8YFzTtPJUpTEffPkfAXvl3q/63pwa/EQwm3rEV1e3sr6pqd7pdtrTsQRS6RTSuSxy+RwyvGYLBEhQRqHAmkdR/QGFYpHV5BNg8t6q1rO0KZSKTo2AWHUduk2H3WY3+Rhly6m66tpdD956xy+/8MTjkQ+GZL35b4F9Z9PHtk/kC9+tq/cvv+PhHVi5ZQOqanxAykBqYBSxiQlEEzHEkgSZSSOZlZpBNkegRo6CpMQIKMerBZgS5f3cSgmTGXmTle8LpgkiVAxJcz4Zn+VVL5j9C6obv/mrfbt/S+Kn+TEL6YcC+/qmB77qdrif3Lj6Jvu2z34Sea8NweFhTIYnkEglkRIJZbOUlACglAzDIl4IK5jqWdqFWHk25UqCc+yXpzTzeWmfBiOAsuwvkk9nMhajyKwM1xAJ22w28IKiYZjNbt/f7T118HuU5HXBfSCwr23a/hc+Z+XTd/feqvV+4n5cGLuCsZFRJJJJSsMCo1SPi2YpGSUNEqxUUIG0JDQNgkAISgBJm0isINIhMAGrAPJe5pH3otrqHRmiJMl3whghWEA6eV3SNG/nvpOH/22WsNTjdYHtXLFxsbvOf2h104K6j335MZy8fBHBYBDJdFqpmCyuwBGUkogQRHCpdIbczpDb5Lioj1JFcSK0txIQBYIEinWZptgaJUlSFMF8Vqoo0hTAUkugLSbkYRJokWMoKTRUVsVuX9a78am3Xjo1G1zZJc1o1+vrvlXI5OrWbNuMowNnEbh8GfFUypIM1S6rOE6u8l4AiFomEpaN5flOpGISiEmiVCUhiuvktPwJxwsEUVBPIgHpxzdsl/syIHWv5rAACiirr7CiiHAi5jtx5cK3Oe7P6HBmqOQcYH+1amtbOJl4eFVHJ2KU9/nzFxCNx5QjEAko1aIEkgQzFYupq0GABbp7EJBOaTjJaSIhxSRdgaBdkeAs7wlXqaTykhSTkCjg+UpJUYHjM9hf7vlS3ctVJFUkAwSU9cpEcGrygS/c8+BiNl5gvVrmAEt7HXckE9GalsULcCEwiHA0gqQYMIkX4xa7ikQmESdYkU6RgHSjAAfBuLi4m9VDzrpJgBu6Mnhx+mmCjhZyiNB2oiQwTQabBKbYrGilc+CTAL0KitTrnEekVG4TacufhAQpKSPn7R8ZuZu3Hw4sWszeLIRoLidGx0N04/R+Yi8EJu48OD6OFKVlATLhNARIEXUkzk8i5js9WFTTgLZ6P8NCPTS3E4ZWUGMj8SguhcdwcjKI08koQgSb0zU4GLMcZJDG8ZqQTeKFbMEjomEz59AxSWbEirRPYQQ7CDZR8ZSRX8+WH7NeLXMllst2aHYbEgQTTcRpW/SCVMEkbWwkFESaAVlA2XIFeIwifAUdrQUbVtq9WNfSju7FS1Hd2IgCHUksEkFkKITJfAIJEmWvcOOGji70tC7AyeAVvDpyAUNGBrc3tmHb5ttRXVfL/EODTsKVSopKUwN0tkUZJ3+2+0W8Pj7EzEbAExUZIhh1m6ON/SnEaTubAUxe3rZ6g9vOARJwpSYIROLIcHCM4BIwc3nYaT+uXBE1BTuWFBy4xd2AzavWoGHdSqQuD+HA71/CuaGLGDXjmEQOcbqJOJUsTQIdDi+6G+dhGaXZW1VLGxnHpYF+ZNrmwTPViIORsZKLFEmJl6RGEFovxyzSbUq905xLhMmUixwA3N4qD3axE628LLIZwKRvOp/THZzAApWk18sqlYzGp+ggctBoTzb6CU/epiR1i6sBd2/eiootvbj00st4/o3nEUACDfYadNR1YoW3BkWnk7aVxvnYGE7ERvHW8HnUJlsRsucwyfBwOkcm7n0VH+/oxjtDlzCaS0IIy7KOs8r9929/wLIrghU1FETi8tVdkQTPKjOB7YKWM/KaqRcQowqmszT0eBwTU5PMInJKBR1UP2dBQ73pQE/Bi1t7VqHiwS0YfO5FPPXGr8hJHR9vuhE9LV1o8NbD9FXhfN8hqqKBG5sW4c7FGzCeDOLc1Dm8HgyhQJt0ka4DtLfKkQFsX9GL508cwgSdjEG6HcRRb7PDaXOIuZWK3BAWjUzuptvL7y1mTD89Qi59yyAJQJxqKB5QvKJcxZ3rooIUtpcqWG/asKKyEf47NiBxcQC/3v0M7cOBT3duwpSRRaDvbfju/3MM1RVw4cQoBjNjmIhxfNGN+qpK7IuGMEU7qSWBNZoNzRUVuH/9LThy7gzOEJRwPEWqK3mlC7LSKaGUKBQ+gpZEWemkciPycrrMlNjevRozcE08jajgFJ2HxDCJXRKXtDzlYepwk6CqvIbmlkZo7Q04/L3vYhQpfIaS8pPo35zYg3p/I370+vdxLBVDnETUOlxY5KnEfK8LFzOTaGhoRRe1IkON6Kltwme33IE9Rw9gz8glLNYdtEwTSUqqgkyqJ70OaoI4EatYzkOjpCW2SR4yu8wExrfMCDTpLKmSCswiLaqJgLLRYrkFU15LdNxe54F5uh8HLh5HFVyYX+nD5MBR9Ppb8J34EN6jsxFyxIOFcgbO5iZQPxXBjdV12NTQgfsbWnCZIWXj0h68ceRP+D2znJscVdQjHUnaZU99ExyxKFITY3Bxn8bIIMsqG5Nbnc5DYEk8nF1mAbtNxjGcFFWmIQ5EpUeUoARKmcGau4gk7TA4OYzwrqOMR2lUMy2diA7jneQwvaEDoxlDtdk1J5fWVcaRZQyK0nG9Tpu9+P5+bG/vxOrOFXj1wNvYN3Aad3uaUHTYEPE6aY8t2HX+OLY2tmOluxNO2YCSeqWGckf7EokpR1iO1tegmwXMUlnJ1eJ086lMSmUbRbpcTYJ/UThEF8xwES7mcHH0EqYyEVTrdkyaOfwxGkDP/B6s712DmpPH8N7IOCZyJhIcGyeoGPskTW5HkMF5jp+4fAF/YoLdQqY8Vt+NOANYwGPDNrr+n73/NvYzlg4HzuEvO5ZiIaUjTkLcvFXEeRCY5UDE1c8oM4FRYBIYJY8T55Gj+kg6oxEYNZSQrJRH8r6IaWCYtdbrRrfmxgG68b5CCs2xCG6pWoDGDhvurEkiVnAhZtgxToqGcmlcySQwkAphIDlEqQZxyExik+aAqTsR9ADbOpfiJ/tfwx6CEmoDrM8GzsJMFTDkcDAEWM5DJKfiGKVH2kSYM8pMYHxVZOIqWwTJ2OUqbBKwkg0It0RaMrkks6cYb77e1o3aKWuTeTgdxisTAxj7zU+w1r8OHo+fTqNSqaTf7kSH04Z4lQOT3PKH6fkCmTAOjR/Fe5GTGB6/hM+03oBnjhzEK5kYvWI5NJm4yPVeiIxijGqcp6EpZpMe5RUJ6Tq4lFedgVQRT1UsSOpEyXBOBU7aZRdlsC1DvXRwgXPxCEaYC9x413bUjATgfvtlHI6P4Z34FRyPT6DV3YZKZyNTHh/sWg00WyVMmxdpgky6mB86PNjccieW19+MfaN/wI+DFxGh1J2aRzSMS0teaNWThQQc9JJWFkli+F4B442QOLvMlRh9onhE6S2SksofDtcoJQ05SixNcAJUAv7TZ97Fmps2Yf499+G2aBwLwqMIjkUwGUvTRnVEM+MIkddBEjnGGaIcV9BcqHa20Isug9ulIa3nsaHpfrxvexdTk8dh08XZGxzBTIfqLoTrXFfceokc3gk2y7SoV9ouPl9b5gDjS/oKay90LbCrEuP0WTFkTis63mek8Pf//kM8fvNDaBocQdgWQU9XL7OOLoaGaqRTBEd7GaVqn58K41hsCEdzAQxkh9CXHUSLoxNdvpvR6mvHYxtuxy+PPY3+yNukIk4QXIeiU46L69LMFaOVhBS/hSqWuSY2VxVpTaqvAkUXL3/CHbnKjpfuhD0sJyIvNHrEvekIzrz5c9zt78Id9k7UhObBt+Q+6Gsb4Lmch7c/zZRoEsWmEOrdBjaQIW/3n8Cey28hYFxGKJLE/VVLsWVNLzIVJn76Vhzh1HFyWCQkKwujuZlS9wLLokvtzaRNIVVkX/2ZIzH2oTKWOTNzkARCSZ8FXJHEERWvBEy3O0j1/Pn4WbymD2GtbwirAwG0xrvhcTQzdPthN6oxMe5gwhtBoDiF9vnrsKP1Jvz6/Z8iYpzC7iu7sPTocuzY0Yt3+rciMjDEozauxCrnpJZPFmcmbLaK+AAFc7qp/GquxEizxY8SOBkpg0XcchVg0iJKoMt+iO3MiUkAzwqpOpdoE6PR89gTHWB0krMkB8c4mLJUwONuRHP1MrTP24b9kwZ3AE5sWfnX+N2JHyJqjuLF4+9hw/plWLpoLY6NvsOQE+MiGUokx3kEVEn1lA5Z0uSLEk1yN13mSEyGXrUtuWctM0SUUo5l6eYIiNmbBEjhJUEx51KHMBkO4FaNQViUWvbi9LAMyIUCd2XM6k8nz0IfOYh1t3wd8XwbOvQWNLU+hMDYf2AwHcCTvz0Pe40TXl8H4pl+EhDn3GSgiqMWPVxUFevEircqAyk1li6WQZUe9u69iuGaXqLPMiEBiHRoUwW7G3lXBQy3j5X7rYoGOHwt8FYxh3DVoUiXbhB0jv0NJVFdqaupkY+qbRwH338GRq0d5xJpNMxfD5NpU5LxrT86gqF8hLttzmtzk2huVziXZelClmiOxWp1FiIt1+ycpYeUORJTIOSHxbpYgJQqEpRGUAVPNUyvH5753fDWNcLhdOHGhcuxsFiFQiCE6JVhDI0NIDBxAZH0GFVUbITOmeNB9Sww9pk2F7wVDPbz3JjkuQo8LdC8tTCrIkjZHcjbRDuuTaGEIgFkUSW/coAq0qQDs5BKl1KZAWz8NtHCkh6X5pgGR7UjMUVKCf4lqL/zISx7YBvCXh8u/fRZHDh5An0VnXBdPAI/g+r8eZ1YtnAjplIZDDJZHo7znD+fgcEgXeFbiBVrbsfDD/Tg0PEUXnj5GPFWoKZrHTeWQbi5+YyE3qMdW39lb2jpTpl0AudpGBi0TdmXzSozgFnvyAGqi+i0KrwXaUG47fRyY9WOuvsewebPPQRfuwd9T7+G1B9fQooSCslZImPWGe7ZdM2L6spmzGvoRkfjSiwkSNNVD8NTj9qFLahs8uLUJLD31CAD+QjQuAhd65dgcXs3Rk704/SRQWYdadIhAZoArvXpijS2UmKa7KzpyB95pCTKEsAZwE7vtZTX4hDnEkCqiRwhsCLtyb2kF0u3rCGj4njhsX9C7NhJHrF5UGxYCB5LcRCpzfLQhy49mtEwRRd/Lp6D1zMEX2U73DXLSYwf/k7gSiiMoaGz3CL70b5lLXy1BuraKvD+K4fJnxD3f+IRJX6Vz4xLMrPUSAGz03HotEP8QwnR9YBtYeO/lizxWgYpbALS5kTlgk6c3/0mDo4EoXetht6wHPbBC2ij3TXX+JGNMbgGhzBy+QzyU8Mo5sPIZXls7eDJV8YGk9i1mhSCx+oQv0Q3vmA+/Dd1Y20jcPDZP2JidDFOH3qVSpKkqvE8iuGDbpdVpDZNvIq1zJBAB6M5KqZfXA+YtNFmWeTn2llkXmZqzDDCu36I2ge/hI4770XiXD/cCR3BIR7TeYoYIdGedBRL53Xjhu67MZacwmWeHU6EBphU020bw1yAqdJoknGLkltzA7pvbcXqCg3vvhCCZ9kKnHn9SaQYB202qqBpqaICpui5hiZSKJmH5qqCTk2aXWaoonrJbFNsTKpVZDLZkzFWpUg5bSj28lPInh1GXd0yZJZuRea4G5kre+Gu4nesqfPov9CHirpVmN95J9avXgkn150wkmqrkqCxo6EWNQtqMd+vwzgexv7WBtTe24DIk/+M8b6XoHm4P6O0UMhwaZGYZWNKEZUqkTaSJbeapxY6t0ezdXEGMJ5e07y0vE69lUT4quRkBgEm7pUf3wrhAL+D8UhtWS+ygVOoengzsr86Bi20l2eOBpyVi+Hq7kKmaxHOaLXIB9KoNFyoaiiicbUP/p4KHlWbOPg8v4ReAOpWBDC6718w/u4LYP5FQFRR7ratynXZV+zeKgRFpksskx20s6aNQasmgye+Ue6gus0ARs9ifvPb3qkkT6YEmGxUFGeUGlhANRVkmQyHDiPBc/vq3scQ+8WXYVzaxwlT1tqxM0iOH4ExsBmtdOsV85fzFLgFIybVpsaBppYivPzU62oehvfoywj/+EUY8UHJvrgmw4owkMDEcVhqqDYspEJoF01iJfM1hoXa1h54bY4JtnwwMPEbazZsPh2fin1MGGRkOLn0V0PKEuSiEj94+piP8GDzjXNciG0az21VYkzCmG4VeRocvfgmogOHSTCPeqpr4fI3Qz+xCL9vXoxcJEQVfg3F4CmCYIB2lNSfeaF4wasqWHL1IjGxDpEU3ary0s6aVizoXglHOHzG4uj07wyJSXNH+8I/TIQnv+XkR4Zsgl9V2CZh0lIFebKkqPReFuXpvLJHCZJcUJ1DqKxfuCrZN52GyfMTBulclK594ABBMFUyyQgjwTG0IxVfS2soD0hg6koYiqlc5ioodmbskvDTtuo2tPHgh5+09rDHjKKmvLblG49/aX9NTfWphvpWVDFfk3NBmxCsOnEVEaVwj8/Kv8guUK0uFFC+ytAFvBg9d8Dc5Ki9Nz8l8QCdbQSTC5M/U+zPPmos5xC1E0ehqoyn+slaak7RG64okiIo2QM6aFvrt3+KH0hypxa69HfYcUaZA2zjxo3pzkWd33dQlisWr+HZhpNab+OZBQ8s+XeVg6VpqEC8E7BCnFWVbQiBBUpFeTZeKSFNpFSu4iDkPW1JIxM0ySfLNlVinppX1imruIBiLJV4uurBL6KZn5+y8cR3f/CDR+lCZ5Y5wOT188/94pnKCtfvfBUe3LV6K9zMEQWcg7ZjUwAFjgy15Hh1ymvAKVVSnrQMkGdbAoT5IvIlgMr7SbxiZiFVjZfZRFLyK96P68i2RABJhsHadc/ncNuOTyPcf/G3O3pCz6rOs36uC4w2k79xWc8XY6nQ4eb6WvzN9sfQxVxOLzJ9oTrYOLlcy+d6zL1K01qqU5YcxaGkqFIi5elE3UTNylexpZK0ZxAm85E0kZSAkjyV9w6axsodX8H2z/8tgif6DmeGh3c++uijnGRu4ajrl/1H9ifvvWnzf54O9HdWeSt7Pn/PJ5lRdPK/PPB7NDkteaTOmCZHYLKTFgeiswpJcs9/qljP0sYq70QOck88VjTitdTfuspY2YpwbgZzm9MNd2UNc+TVWHvfV9C5aiuCp87uDvUd/tQrv9gZslaZ+ytrfWgpPvGEftfv9n7KY/d8rbN1Ua/H5dWG+dUxEOK5fWwSyVxK/WcT+b8feUpDjscl1ZFNoBKGWkGILcMqL2ctLTJW7FDghEF0VtyP6fScTncVKqr88PkXoLKhnbum2qPZePJHB198/GkOu66kZs5efvqQ686dOx197/aty5mFDQVdW5zJ56p51kGfyPN8qpYQaPAbmvy3I1EjsQ8BqLSRLRoPWHXNTlOiLVFUYgOSA5SkxTMhK0sV1WM/hkIX/3lMt9eXsFfUDjLZPXh6z6uH+J8DaKAflY848BEH/q858F+Q/tY3+Zi9TQAAAABJRU5ErkJggg=="
_LOGO_B64_70 = "iVBORw0KGgoAAAANSUhEUgAAAEYAAABGCAYAAABxLuKEAAAAAXNSR0IArs4c6QAAAHhlWElmTU0AKgAAAAgABAEaAAUAAAABAAAAPgEbAAUAAAABAAAARgEoAAMAAAABAAIAAIdpAAQAAAABAAAATgAAAAAAAABIAAAAAQAAAEgAAAABAAOgAQADAAAAAQABAACgAgAEAAAAAQAAAEagAwAEAAAAAQAAAEYAAAAAXao0LwAAAAlwSFlzAAALEwAACxMBAJqcGAAAI5hJREFUeAHte2eUXMd15te5pyfnBEzCzCAMMMgAQYIkSIKZEklZK1HmUjS9VljJ9jmrddAer7z0OXvMH7asI51drySb1DqAosW1mKwlRVJMIAJJZAxAxBlMjj09PT2d035fve7BgGgG7Z794XNYwOuqV69e1b1f3br31n01wKfpUwQ+ReBTBP7/IWD7f+k6e+iQ68L+/op9hw/UBbKJsmg6bU8hyX9AIpG0IZVEKpVCKhpDlGXesCqFTDZri7EMtnTyN6Yik+pVkQVsaWbpdDprt9mycDrNMydcMC/wx+1zZZ2sh9PFchHLLtjttqzLWxRZ17l6+ut/8h9mbDZb3Or51//9vwJmz33f2HAqEvj8RHD25mQ63el0uSrgcrgSmRQSZDiRJgAsk1FeGSQJSjpXTmUyyPBKZcg6E4kXEKZt1pTyZSLDd6w6tcmyrT3X4vJ7Nrvd9KF+dNlt9jTBDLmczsG6mprXt67d8LPvfO/P3+Uza0Az6sf//FrAfG/T3av9FcV/2j8zeX9RWaln2407sen6HWhe0YoidxEyl2YQHpnA/HwQoWgY85EFhCMRhONRRJJxRBJxxJIJJCg1iTSliQClCZzyJIFSnuJ9JldOZ7JIm3LagJnRfTaNdDpjgDLAsy4PmsvlgsvtMuBFYzGEFkKIR6Pp2vLKX9y6Ycd/+aO//vNjHw+J1eITA/PtHXf/dsZX9BcL0XjVbbt24/YHvwBvewMykTAiwRCiC2HE5nkRiHg8hlicIPBKUXqSlCLr4jLTckqlrZzPJFFpAcI6tU2lxbieaxla5Xwf5vkH2+Xe0XvqOxKLIkhAVHa73bA7HFhYWIAjlZnvrl/27afe+JcfUnokpB+ZPhEwv7v1tm9X19Y+Vpxy4qF/+wgadm3C3PwsZiYmMR+YQyQaJQgEI5FAPJmkRCQMYwaEHMF5Ri8DIYC05KylpXotMS0fSYne1b1A05V/JnDyQKrPOMfSuGZ8I41ctnxf7bngDPAOu8NIUYYTtaKy4bFfHn37Tz4OnI8F5utbbvlaU9PyH5ZFsvh3X/kabD3LMDQ0iJmZGSyEKR1cHkmBQSITykmkchFN/UPmOfPKBZCYypXzEmIxbi2PPGhpttNyMyBIovjOIkhcRtJf+f4ErkAUGAZA8zy3HLXk+D51jtFh6t9NWVm/fMUfvHj47e9+lMh8JDAPdGzc0rFqzZuZmZDvmw9/Bd7t3Th75iwCc3PUHWGCIskgYWRE4Gj2jJTkiTUgiHCBYylhvSMwBWSSzOtdIzkGNEmHxaCZdaOwLQUu/WISmSS3lm5i/wJt8R0BaIC8LGWmH9ZLMWclRcxrS0rjO1ZtvO3J155/68PAkbUsmHbRENa3tvyF3+/33bFuK1y9rThx7Dim/ZSUqCQlB4oYJBCG0YSkJDd7JCbO+wjb6ooJDEmTQGL7/NJYlBISrWUkhZq3RsJCzwWJqVdJ/3PtDCACQ0tNgJicfSj/4KV+WOdwODEe8HvOjw781euP/mTnTY8+EisEwIcCs6J35+6Y3bbLHk9i7e03ou/cGYyOj2E+vGCYNJJilszlWdfsSb8YS2QAiRrpUH2WkiDC6JwgkxNxiXlWViVHdEqgkEr9CgjLxOfrcuCwrZ6poWmjfgWUgFCu5xrnA2VWGP5TpM9JpXxpamzzf3v1yftZ+dNfCxiUl/3O2NQk7t52PaaTUQwODWEuNI8IlWx+1o0eoZUQY/F4AvO0BgsETtKUllQIEEoQpxIOEmsjEyrzAS8pRzGUNSY6ScLlpQgMlWXGyR4v6gflwoJ1Ys+SKCsnsuaeP6rgGHyDudroWf65uRcCXIl2ZiFar6Hpyd/JPpr9J9ujNhFzRSooMV9t3dy4gMzNMVqbljXd6B+8BH8wgDA7kwk2OoWMSxKkBOcJWDA4z6USN2BkCBYfwM5nLhIqUNzkzEViVbYTHDvBsRig/0vfJE6WoyTa5LyPkqkE2xv4pB8MJKSddZaxFUS6Z19qx1t7HiS1ZZ2RUOa8sYBiyU6HUErdziU1E5rb8fV9n1/B6vO8rkgFgYnUl22aiy5UlpWXIemwY3JmCiEqW+kJmUYtF820JGNm1m/8BCMdBMMmQFIZOMm8l0T5eBWRtmK2L2Ne4XCjssiHEl4ujxsZTl+SYEVTcfjpFE4Q5LHYAqZYt8BnUTKVJDAGICMBkizZDDEr/i8zbiSJ95IrOYfWpoOgsk6tBF4+0UOmwxkrmoqEtrHukwGTcDnXyz9paG5FkEtDekWOk3wUgSNTK79hcmba+DAZSg/FKAcIJSOVRTGJqCADVdz+NJCV9rJKtFc2oK6sCiVVVXAUFyPrdCBhl3OXRJT9hwnMXGgOo/SRLgSn0Tc7iYE4lbeDTJH6cq8PvjSZpDVz2S3rJIAMTBxPEphDi4Bz4tj/cCKKWU6StVYsyZEC1rJMsq9YMrkmD9bSvKDEJNLpZQLA4/MhQPc+TBCkWyQtsihaTuPTk4hycyhQsrwcSS4ZAuLhKirLOFALB9ozTqwurUBPcxuaa5vgbWmC9jbJ+QUsTE1jnjrMHwsiaKNVc9lhLyuBr6ICK8sq0FReg2WlVdg/PYxjwRnYuWG8ta4Ju66/CbUtLWZZyj8h/yYZ5St9RgRUL2UyMjWGx59+Cr8KTHK7SiliraRHNGiZ2ej42ZzuRquHK38LApN12Irkk9i4jObpTkcIQCRGc8s6OW8T01PGjxFIRMpIiispUGyoTDuwjLvg7owH25tasaazC0VrOo3/MHPwEPqPH8FwcBKEFQEK+wLJjZCbGAnXRS2E2spqOWGoKylGF3fNU8Wl8HOiJs7QMhK4Vl8JRuamkMgxKiYlKXkTneSaafKWIjs/hzYy7yNIYYJlJIrvCBgtfTufFflKi6+ExLorCAytgkPSISUbjlGUuQmMUmKktKbox8wtzDN8wDmgSNspji5ejDeglBLSkHWiJ+vDTR2r0N67Bo7tPYgdPY39nLl3Q4OYJhC0VYTODrfNDYfTS4XMpZmOYYaql7scXAyM43RgCp8tb0JtlddI5QT1zoXkPGYOvI5i7s32jgzhudlRY2FkZSQ4kgiBFeb1UF0LPrPtWpasJWbJkNVOu/TFlLEVxKBgJV15m0IH2gnLmdNOVaZZmzP/HOeZz9IERpbHSWrsvIooKVVZB1rTHlzT1IaObZth370Z8//yOp5/6nEcTs0awlscpVjbsBKtda2ocPmQctio4LnRyyQwseDHidlhHJkZoD6jhfMC7/lHcJY6Qs5biDP/Funwvn8Yd6zcAA/1zCX/GHm0FK0kbpp3g0TJwZ22QwDkFG8eCKkhk4iUvGHO5xKU8q1yYZ/Lt1aJzptN+xxJjZugyG+RFzsTmDVuv/Yy3LAwmqRLS8iO8qwd9Rk3el2V6F67BvbbtyC6/xieefJvcCwzjyq4sbOsA+t7dsAdTcDLf6n6WkyfOo+j4wcx4LGhuqwam+s7cfeqG+iAncLphX68Mh1E2ltClmnhuAQSlK7XOWHlY/3YvnYr3ts/hfEUrSRJT5FZUoYSXl6CqJ11HgeBZ0DJAWJiOzQMmYxk6er0oRKjZRQhKE6Px7j8Wj5hSo883gyf2YxJll9io1m2oZTSUp9yoLutHZ4tKynTKbz++I9xJBPAchTh3uWb0VzZgv0nX0M37XDLN/4Q6XaKRGAY9lCW8ZkoDvn9ODp5Fg3FtfCWu7HfPwo/+64i3VSTKLWTXAJzbWsH7undhmcOvIULBEWsCYAEf4qYq72XIFKZWMAsiolaKaBlPdO+SXLDn6tSQWBS2TR9MHqzFGdZIFmo2eCcMdfyaOXeS1qcnCYnJcXL8XwEqJr6onp5PdDThvEf78He+Ysoo6TsrulGV9da7Dn0LM5Gx1C6fDVeeOa/EogRDGZisLs8aPIUo7minHrHieEwNc3IPMa9dvjcXjOOj6CU0LPbuXI1Pr/jevzdi8/ilZkx9k8rSEa14ZmkCS5nXsnLyzr5O3nQWLWY8lE/U1EYFyuCuvgGC+TRdg0vueRaTklG2qRn5GNIz0ibWy6+3G/LXGo2HVxSvhIPXI0k1R/EobdfpjLN4BqbFyuXd2Fh4H045sdobTrw3cA57KMERjmeRF8m38b+i2lputwebKxuRHF1C+6tqgX3azg0NIAkd/Q3r1mPz+zchSeeeRIvTQ6jyeZAGycjyjaDyRhur2nE6PQ4qjluMWlyUmpkuhdFwhKYnMTImyYreVPF8tJUWGK48GTvzSaNQMhky7mTbslKr9Ac2uVGUloc6p1DyN9Mu0kCe0zsfRcnFsZZzy2+ywt7YgHj46exrrIZP3Um8QpBkVtGWM0SkfaTpVLtYeq1c+OXsJHO3HanHT2tK7FmXTUWKL07enrxxLNP4sWxAbTZPFhD3RNmG1ms+zrXIZ5NYUtZDfovHkcxQVGAymwfcoBwGJO0LRBtuuTVWLVX/hYAhr3YtxlgtJzoGZrtQD7mIZ/BAoXd5tauxo2TtaA9gUsn38XkuXMIMPYsolJZ+j3+YTyfYLiCDuDplIeLywYPGRMsGdIlkJKSUO6RqMHo22Swl/pt6MwxTE2Oo7drHToaW/DEP+/BK2MXsdJRgq2+SgRcNpyJzuHBzTtxbnIIr106h/s61mBtRy9quPSy9MPy0iIa8+rERkkzckRpMvVXYmLuCgBjKTJpeRNbof9ifBgpXYKiDYcYFiZGqnhDOaIDlaa5ncOF8Sj6I1PUBw7McQYHkyF4uYSWlTWifPVqVExcIn3FmFyI0mPhho4XDT8S7CdGcCmXBJlxHQI6wPJcYBrD777NVhk6gllsdVfjhrJ6jNJxOhf246FrboJ/5CJeGDiLi+rn4il8sW0lWtpXwePg5xYm8bIUAEv5EjILmE8oMewh02v6o5vC+AqVr7Vp5NaMwEhKjKBwGWkwLSLNcoTWYzA+j9WpKkzS9K60l2EiOoNxMhmOzVJfrMO6mz+L3rlpXHfsNC5M+zEcSWI2kUaA+mmW+muGlimQiiCYjlC6whwrhjm+/yb7X8Y53u0sxRfqu3HGnsTJuXE8dMOdCAyex0/O9+EsaeGWCmd4/fzSWaS4JJtD7Ugx9iLgLVpZkAxJ7/BSJdXFJwSGr+oThaRB334WqBTNzlmASGLYmURfgGgw8mQkhrsmDJM5D13xyoQf7aUNGB4N0vXnzpn97Zs4j6qXXkdVUzum56PYVLMc2/k5KpUm4bYiBNnnDN/3xxLc+IXRH5pEf3gUI7RiM7Fp9GeieCkVhW92gm2TuGPrDiQnp/HD04dxjO8KFGvZZHGS5eaxYXhSFejj+DGGWzilFu2UTMvzzS2jwirmaqvEPg0AAkY+S5QxGRMREwwSFeu/GYgq2oCiGIr+RdMJnM8kcWfTKvgCcdzYtBqvjfVhgpNyLBXC+NHnsGtiOzeHK5EKeREj8KlshP4HlysJdHHXW2F3o6y4Gd3lKxDj9mKK752YPYOj/uMYWhjE34Sn8EV3M+YHA3hq8DAOkRInF5p0hgyAlbLYR9rOTb+PCVFIXSNDYqaSkm0tJbak1FBjWnjm3sxnBXWM4Z7N5cMocG26FChW1yRD65agsC5FAhIsC6QsrcDr4/34fOdvoIVmu9HNdgT3nakLGOBb/QRtaHQvGm2n0FLShqqiJnhcVXDZyqgP6K9SITsJTmAuiWlbEFNUw3G3Hd2Vq9FdvQknZ/rwzvTbeCY0hr1jmgR65lLlYk0CQQWuPbTyEKU0xLKbeiavY8SBaSprlZMU4lQwFQSGPoz6MDEXRek4mllaeYnRQIYEgUIinKTM4EYzOBYP45nTh/Ctr34LntZl2FpWhupj76Lv9EnuYag/CJDClxdDw3g/NErWZdEYNOJiiHPu0wTH66hAmace9QTPZ6/l9oAQZcJYXt6NqpZWHJj6Fc5MHyKgPlpQTZuMPWPJRn5FueixZEGGQbQZhvSIyegXq6Q9WEFoCgJjRmPnitCZrbxmwsJKCJkZoGNkJCbJPM466R2NYLe58MzYabQ9vQef+9rv8jsOg0t8f/3u27DNn0Z4eBbhIL8ecBs8z578ZGiKOmOKTI3RLk1lExhLjSOYGkQq/B692lqsLdmOBi6tU5PnGcj2YX3NnVx2dTjv10KKsgdpuCSVNZcUrZlAEjGWLjQImB8BZKTrcpXhZcntYrEgMNmsVKo0tuRCiJvfHOwq65k1T/mlJJE0q5X5DAn8wanXMfX9EP5N/QakhqfxTv9JdLZvQseGbVTMy+HJUvHGXIjH6FlHGNaco8KlaT4dHEVffAT9hGqWi2E2O4dXQi+iPtqGrVV3Iety465196JpqhO4UI4LM/tITYi0cW+dpeyJOF5mHjlpSov0W4+se/GkmaT+KZQKApOhD2KmPw+KOuGlpL4smHLAEDyzXq3HHMxGn4Smm6P+4OJBKt9zuLdhPba51qIpyLDmbA+y6zfA1kXix+mxTImhGBwNMTSvakdjkRvXJu14f2QYL/ftR1+4j33OUIpG8OrMs7ih5iG47UX40h27MPt8BsFkENOh06RKUTn2afSviLGuRVAMfRaRYsXoo1zOxlelwsCQfwOEeVE/ufeUExllWtWyAxRak5tHeiZ9Y7xZ+TYO7KNnemLgTfRQZ+ys3YhN7fVorSzlTrkF3vpKlMa8KBnleZcLIRwJ9WFv4hD8xUm0r+7Ab973Fbx58hx+0ffPBPs85ik9e2dfQNF7y/DoNbfi3ts3o/+pCQRidAqo5O10+kSRjZ9wZW8s6kgZJ8mScxKrW62EnGpQaL1QKghMjlf1YZKREHMjybF0iSQnQxC0CTTPWWHnEjKgcKw0w/9abk6SyCAi3osHKAVvonrkIOoYJi/jHtjhKYPTx0hNcTVKSppQWb8cbb71qE54cKD/Al488RLuvO4O3FX6+/jZuz+iN8zPw9kZ7B97Cy+/twGfvaceXS29uDB/mLHpCAGh4rLRneMO3QBjpDzPhShmWZO3WGVa6cFVqSAwxpEzHQgI61pEyaxbSQUH4LWo0wmStZMlOAYgtaECp2wzHGxCjgo7zhOmYWObaI/iDqTiVOIBAcx9E02vi6A1VK7D1o1fQpFjGx4/sBdfvOYubOx6GPv6/5YDjvDTyihePTmENV11aKhrRnlpN+bCQySIfeoQlU0ORI5A5VKxujW4kK4cT8KHn1nyMF0BTkFgaOgMDlYHV7Rn77zXWPwR48ajpF/AyDnLDoIjYNiIn0XAGLC+PgpofkSgFZMmkPwol6hrMyGy2Rdz7ZJ42gazgQD6XjuHbTt+DytvvB2/ODSIrWs30tG7GbPz/5ueNHfggSH85XPlcJba4KMyt83QD8p6eJEl6UjNWq5fA5K5M4RTD8nX4YjyZRjlK5RE49VJtXxPxOaTymbJmFxlMUNgCEqaTlTa6WHYoQgpj495Me99yDh4yop+SYZB7ww92iyvDM15moTrI5r2MFqK0gxW2YIqY6NCtk1j/zt7kC3nV4qyNvjjKTS27UDG3ca+vdxDxTHIoygjkTk4ShmecjLY7yAwdm0cNbG8LByEj0l5fmRtrWSBl7u5IisoMWYaTbN8V7rJlcmQpWDlWtODFDGMuaRdDCryyhIQF4NNcu8rS2tQ7ymFi0xF/HP8jEtHLepnbCe3/zJ0SWfl/rFvE4qgZpJ2Qprm+9QJhug24WI4CkdFJWxFLUjzoEGSExBj6DviSKLBU8tJEBCXabPQyDG+lA1xQgmWxBhHr7DAaPQCiRJj9ZWDOtfEEkkxwQaUlCw/t2YJhs1XgYyvGu7mLtSs3cBBKUn8gtDZ3IK1jPOWL3D+xmaQnuK5vJlZDI1fwoD/LCYjQ4ik+CmGFkJE2qk0s5SoLKUMNm0RatDR2YnG+lZcGkpgaGqEwOtZA4oampANn+K9l9LHrxbcAkgOzPLPs2SYsDi5zA9pEzBsow9udkpzoVQQGEXpzBrkG+rAuggSiVeS5ZHI2igp8JUjXdmKyu03Ydmdt6Gxox4XYw6M/N3Psf+Fn+LQ8hvhCVP7DBxAZTaK1uplaG/txqrODfwck8DF6UEMz49jNj7FTSOPlNgZwHJWoqSsHd0912HnrvVYVuvA0LMBTJ4ZpNqiRiruQsPqLoyeHERNdRvC8xcIrvZ02kWLdjGe/2dIvvxDFkxcKceHBdHlx/lSQWCklASBtQs1pVx7lgWOUVqUFg8/a5Q2onLnbWj5zN24dlMDZms9mHj+NOJnD8E2dgGxwbM0ErRHDHhN0qE+03+Ce6tXUcVl1t6wCp3LtqKzczO3FUUIMMy/QIXtrq5BU2c9mlp8cBelsb8/g77+c/ySMAt+N0bpil60dpWhoulOlLDPfc+/SDAUBSA4BPcyJJaciHhxkb9bqnx5OCFfnePRygoDY5pKuS55Rz0rGWnhaxThbFEVinu2oGH3Dbjl2iYujQhe/YMfI/LzpwEGkjRz+tyRZZyFkSNSpsVIE23nJ9ekD35+GzoyOoISdzma6jrRUL8SxZUdlEQXJudsOBdJId7oxKmzfgT5IY7fRIDaLqzc0YKKBmDr6lIcfOYM/NMDHCdM3aHzznI5Oa5RsKI/f4l261bAmMSlpM/QhVJhYNiBQf0KMPPIWPqFH5xgr2uHr2E5GuxhHDp7CQf+0/eQHB2HbeU1sDNgBR7pyHL/k2XIEyE/EAkgyyBUlo5YJj5JAJJwlaxAlOb2YsqFC8OjyI6l4CxzwLmNm8d7gIWJGQQH3uE5EpJa246Ga3rQ2RTHvn2TKCrvwOmjb/Jcjp/KkgdGKC0ZSQz1jYBZqg4M88JIvFHHyHKBVlKb3kKpIDAO7gbzW3NhYylddcpe8/hIWTLM4OvowoG//Rli0wE4d9yE4gc2InL4OBznTqLG4cWKzbtR7C1DjOeAZwIzGB6ia68ZjviNJCUiAwyfzsDmbaDFqaNZ50cVgoehS3jviQrE50i4l+a4hYp903Lc0uXBkR+9yGVXguPvMO576Jekg8cCGN2TxGS5HeCnDNIs5kV8Dg1zb2pYRT3FYLmNboWdkl8oFQSGPOf2pUuAMINoHA5ExG086JMeOILB/34J9b/1h2jioQF3fRmGnngSzYzN+mdpohkCmK7yYWpkHD7qmZXL27GpYyeGQzH0j17E5MQ5xIMjBGKB/eqQwAQHoHRx04iRBsS5rNDcCduaNu7Ky3FztR1Hf84zM6EWrL6hEUd/+h0koqMMRVC+GSfO0vEj2oZx/izhVwDp1ppVKV8Hj5XYvNyWyIAUSAWB0fEIdWGkxnSY61g4GRGiqMbJTIzb/al+LPxiDzLbbkfofx5H7/1fwmwlQw2PP4HowEsIDg8Bbn6EC53B8QMJVFdT4a65Cdet3sCvNNdjLBLFyMyEdQiJpypS8kRp6VBdDxfP09Svq0VvhxsN/ijeGeA3pO116FlRh4knH8Psxbdg9zFCx0nKMpAO7ZUIjCTGIEFa8+Qv5d1MLiXGzn0a7HQLCqSCwFBRcwskFPJLSuXc2zmFCsY+tI4VzwifOohw/yTabvtjnHjhV4iv8iBbfz1FnH8ZQktS5uQn16CTWwIfFoqX4eCEFwcnnWhujmNNG8MMnd1wbFmLKEMOYZcTMa8D7ioXeG4J7jkejPzHfrxx1Anv73VilTeMkf/1fYy+9Q+MhFI6zNKhchcoDJ0SJRJKugwomlBBo8vix6gF3tqoIx0ldSwUa81dla4ChlKS7ezZHLdx5vJxUWGirq3+WTLKzdQQcc4wPVlkubk7vhc++i3Ooin4vrwV83/tR+rkU/RRJvky7ZGrHN7mG9B8Yw8a1jTzTEw1jg+m8db+EH2deZSXz6F4tQ/V11SimKCMc0EfPMLvTmdqUUnd4nv5PZx640cInnmTVlEfeEm+USUEREsxJy02TVgOHBaY8rOqnBctq516z13RTCNGD7NAugoYtSkpKw84KGo6qyZw6Cty4KViKVAY95Bml0KWsqNliPR9F26Pi3+JshVzf/pVJAbeZjsuN2NCuQ9KBhA/P4bgpb2Yat2Blk3XYu2aHti6lyMYqeMpSg9mW2woWgWab57ho18T3MxPKvxKEN37HMbef43WbIpSSumQ25BfNpIUTZbMMEExlJJewWBRbQEi1aCAvY17O09FE6rKG+CYHtSRmqtSQWBqqqrPDnr4zZkd2G10nHKMWW9LYljSWEpmPRM4ftfWLM0d+R/AkSf5nMTKwohoQxcZoQ9jAelHaPgNnBo5hlMvVnOHXImyhnoUNTaiaLKJeqkFr3BftMDvRrPvvIHY+3uJ7wj7omTqlJKJoWqyBIR2ynm9IkkRfSJQKQ8OiwJFu27FaqhwK5f3oKaiFLY55znT9AM/BYFpW7Hi3ff7jtOndDrizjDSPDREajiOBtZw+YEvEyLxtUykJDOSa0up0gd0SVWOKCsXkXLGaH2SXG4BJ2bnSMo57oMYvEJJLf2WCir4IAM4Q+yOUkLP1kge+xLOlocrcHRZkqKy9T1d9AkgZhrbvCE6KC2UGGdpHd2IG+BLpTK+2upDbHBVYuur0x9//7G+mtq601KelTyCqhiL+SfxzYuKGZtE5cRXVBgSuNb54YVlhhdNBX+0HJVRZ5g/FeKXQf3BGf8YzZIAhhDMyTkeMkJ8mIeJToDalRbvCEGRCefSWRxaEsIx8lKSkxpr0kTUksQxTY3ollrgCtBV1bkdq3o30htPn7l9S8PxJW8sFjXcVamLf0vY1b16j50MNDWuoINGkeM/BxEXQIZfM6RBh6PnchGih3qW84Qs0uSFWjNreaVizLr0ucOAqOVKsIxU6JkxvczNMua70nNG10k6cpcsUH5ipFNydEiiLYrIniRVzpziNJKWsjpsvvsLPLFOr8Dj+vtHHrlJZ46uSgWBUatb7rjzJ/X1dWNpHshZvWIjDxzrQ6idh3Ec1BQUSQMQ4RIFhgzlpMNc+tWDPDMsa6kJHMNMjvE0JcUozhxILJu4rTG9efNr5Qpw56XEJmAWAWGf+YnhiErWYhMovAQKwyM2SozD6cby634DvZs2Y3JgaGRVa9Xj1htX/34oMP/+y5+bWt+74c/iXOfLePC4t40OGf9JapyyWIvSIzIsOKzujchcHklECxABZaRGjIgxzbYFUJbfvBl94j1jtSqnOIm5e6sNTbEBJi8pzE1/Av/KpAVtIv85UOgjGEkRSFXrbsFdD38dgUsTKLKlvvPo17ZQyRVOH+Diykb0EO333fPA348MDD94a8+NeP9cH946c5DxWwWbSTtnzZyM4F3+n2bPkGt6vrJ7Yy4/CCKbiBn9UwTOpHyuGyMNS6TC3Fv1ap0bjaUl7+YlhRMo8yxJqe7Zhfv+6DGeRXZh5PjJf3jqz259WD4bXyyYPlRi1JovZm7ZuvobdTUVL+87/TZu23o9vnXXb6O1fBmPmHFAmj4X165yOyXI8mukaEVkjtAlwxpXfKkECcLF+5wkUaq0VGTl8rkFDjtaysbiECwICAOGpU9M3JeAqM7tK8Oybffh/v/4GGM3Llw6fOyXxbHhb34UKCKZb390evGNN+Kfu+X+54Lh2eYj50+u37JyPb7y2QdRX1qLuWAQC/wjBhPrFxg50ywJIDpmNvV7maEcZ4sTZcmZaWAAUtuly07t83WiU+Vcrswka6yl4Nj1R+o8n1fR1I2e3Y9g+z0PIzY1j6FjJ//JEzr18J4f/H5Bbzffo3L2+skS6bE9cOO9vzU9O/ufW2qWd+zeciM9XC9O8fTSqeHzGObR0jm69fxTF0ba+IdcWma5a/FgAJm3+JfPo5QbfsnSseqtZ4vL5HIlH+iGCy/3jvlDCUqHcUZ5LNZbVM49aB0qG1eirr0XHu6HAtMzQ65k4rHX9nz1x3xZyH9s+sTA5Hv65v1frh6anP7NYDD4gMPh6fV6OTUkVEde53l6MipgGLFLUlmao/UGHEZimeejiMoFkHCxcjGaW4KsNzgsAcsAocas05cJK1eZS5jKVTrEwVMQbp4V9nJn7qFz6PSURDN2Zx+DLk+VZuL/uP/lv6SX+MnTrw1MvmvqC9uD9zzYPh6aXhWKxJfzNFUFfRjSSO+V1ssEnLm0lKd4+CjOk+J2Hj21nnHaOG+mjemQ9VKSdp5V4AN+hzbY8CMUc/5JC/uzLiNvrDR7glwd9YFcCUUUHY6My1sy7ywpHS0qrTnb7tly4emnvyAT9mn6FIFPEfgUgU8R+NeAwP8B9b7EDfdYgZ8AAAAASUVORK5CYII="

# ── Column layout ─────────────────────────────────────────────────────────────
_DATA_COLS    = ["Hook", "Text", "Prompts", "Titel", "Beschreibung", "Fertig"]
_DISPLAY_COLS = ["#"] + _DATA_COLS

_COL_WIDTHS = {
    "#": 6, "Hook": 44, "Text": 48, "Prompts": 48,
    "Titel": 30, "Beschreibung": 48, "Fertig": 14,
}

# Ein Gold-Akzent für alle Spalten-Header. Fertig bleibt Status-Grün —
# einzige Farbe hier, die eine echte Bedeutung trägt.
_COL_ACCENT = {
    "#":            "CFA347",
    "Hook":         "CFA347",
    "Text":         "CFA347",
    "Prompts":      "CFA347",
    "Titel":        "CFA347",
    "Beschreibung": "CFA347",
    "Fertig":       "3ECF6E",
}

# ── Theme palette ─────────────────────────────────────────────────────────────
_BG_HEADER  = "120F09"
_BG_ROW_A   = "17130D"
_BG_ROW_B   = "1C170F"
_TEXT_BODY  = "E7E4DE"
_TEXT_NUM   = "9A9AA4"
_TEXT_HOOK  = "F1EAD8"
_BORDER     = "2A2521"
_BG_FERTIG  = "16301F"
_FG_FERTIG  = "3ECF6E"
_FG_OFFEN   = "635C50"
_TAB_COLOR  = "CFA347"


# ── Public API ────────────────────────────────────────────────────────────────

def save_xlsx(df: pd.DataFrame, path: str | Path) -> None:
    df = _prepare(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Shorts"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _TAB_COLOR

    _write_header(ws)
    _write_rows(ws, df)
    _apply_col_widths(ws)
    _add_status_dropdown(ws, len(df))
    _add_status_conditional_format(ws, len(df))

    ws.freeze_panes = "B2"  # header row + # column always visible
    ws.auto_filter.ref = f"B1:{get_column_letter(len(_DISPLAY_COLS))}1"

    wb.save(path)


def save_csv(df: pd.DataFrame, path: str | Path) -> None:
    df.to_csv(path, index=False, sep=";", encoding="utf-8-sig")


def _auto_tags(hook: str, titel: str, topic: str) -> list[str]:
    _STOP = {
        "der","die","das","ein","eine","einen","einem","einer","ist","sind","hat","haben",
        "als","und","oder","für","mit","in","an","auf","von","zu","ich","du","er","sie",
        "es","wir","ihr","wie","was","nicht","mehr","auch","nur","noch","aber","nach","dann",
        "wenn","dass","man","kann","wird","sich","sein","sogar","schon","immer","wirklich",
        "doch","eben","halt","denn","sehr","viel","viele","diese","dieser","dieses","uns",
        "über","unter","beim","dem","des","den","bereits","immer","jeden","jeder","jedes",
    }

    def _words(text: str) -> list[str]:
        out = []
        for w in text.split():
            w = w.strip("!.,?:-–\"'()🌿🐜🌳🌊⚡😱🦁🔥💡🚀🧠🦋🌍🔬🎬📌📋🏷✓")
            if len(w) >= 4 and w[0].isupper() and w.lower() not in _STOP:
                out.append(w)
        return out

    topic_base = topic.split("–")[0].split("-")[0].strip()
    tags, seen = [], set()

    for w in _words(titel):
        if w.lower() not in seen:
            seen.add(w.lower()); tags.append(f"#{w}"); break

    for w in _words(topic_base):
        if w.lower() not in seen:
            seen.add(w.lower()); tags.append(f"#{w}"); break

    for w in _words(hook):
        if w.lower() not in seen and len(tags) < 3:
            seen.add(w.lower()); tags.append(f"#{w}")

    if "#Shorts" not in tags:
        tags.append("#Shorts")
    return tags[:4]


def save_html(df: pd.DataFrame, path: str | Path, topic: str = "", music_config: dict | None = None) -> None:
    import hashlib as _hashlib
    import json as _json

    df = df.copy()
    count = len(df)
    topic_disp = topic.strip() or "ShortFlow"

    musik_enabled = bool(music_config and music_config.get("enabled"))
    _musik_map: dict[str, str] = {}
    if musik_enabled:
        for t in (music_config or {}).get("tracks", []):
            _musik_map[t.get("energie", "")] = t.get("track", "")

    row_objects = []
    for _, row in df.iterrows():
        energie = str(row.get("EnergiTyp", "wissen")).strip().lower()
        if energie not in ("phonk", "action", "wissen", "clever"):
            energie = "wissen"
        row_objects.append({
            "num":   str(row.get("Short", "")).replace("Short", ""),
            "date":  str(row.get("Datum", "")),
            "tag":   str(row.get("Tag", "")),
            "short": str(row.get("Short", "")),
            "musik": energie,
            "track": _musik_map.get(energie, "") if musik_enabled else "",
            "hook":  str(row.get("Hook", "")),
            "text":  str(row.get("Text", "")),
            "titel": str(row.get("Titel", "")),
            "yt":    str(row.get("YTBeschreibung", row.get("Beschreibung", ""))),
            "ig":    str(row.get("IGBeschreibung", "")),
        })
    rows_js = _json.dumps(row_objects, ensure_ascii=False)

    import html as _html_lib

    _musikLabel_py = {"phonk": "Phonk", "action": "Action", "wissen": "Wissen", "clever": "Clever"}

    def _he(s: str) -> str:
        return _html_lib.escape(str(s), quote=True)

    def _copy_attr(s: str) -> str:
        return _html_lib.escape(str(s), quote=True).replace('\n', '&#10;')

    def _cell(text: str, css: str) -> str:
        return (
            f'<td><div class="{css}">{_he(text)}</div>'
            f'<button class="copy-btn" data-copy="{_copy_attr(text)}">Kopieren</button></td>'
        )

    _tbody_rows = []
    for _r in row_objects:
        _tags = _auto_tags(_r['hook'], _r['titel'], topic)
        _ig_block = _r['titel'] + '\n\n' + _r['ig'] + '\n\n' + ' '.join(_tags)
        _musik_cell = (
            f'<td class="td-musik">'
            f'<span class="musik-badge musik-{_r["musik"]}">'
            f'{_he(_musikLabel_py.get(_r["musik"], _r["musik"]))}</span>'
            + (f'<span class="musik-track">{_he(_r["track"])}</span>' if _r['track'] else '')
            + '</td>'
        )
        _tbody_rows.append(
            '<tr>'
            + f'<td class="td-num">{_he(_r["num"])}</td>'
            + f'<td class="td-date">{_he(_r["date"])}</td>'
            + f'<td class="td-tag">{_he(_r["tag"])}</td>'
            + f'<td class="td-short">{_he(_r["short"])}</td>'
            + _musik_cell
            + _cell(_r['hook'], 'hook-text')
            + _cell(_r['text'], 'text-body')
            + _cell(_r['titel'], 'titel-text')
            + _cell(_r['yt'], 'desc-text')
            + f'<td><div class="desc-text">{_he(_r["ig"])}</div>'
            + f'<button class="copy-btn" data-copy="{_copy_attr(_ig_block)}">Kopieren</button></td>'
            + '<td class="td-status">'
            + f'<button class="status-btn pending" data-short="{_he(_r["short"])}">Ausstehend</button>'
            + '</td>'
            + '</tr>'
        )
    _tbody_html = '\n'.join(_tbody_rows)

    _CSS = (
        "*{box-sizing:border-box;margin:0;padding:0}"
        ":root{"
        "--bg:#0b0b10;--surface:#111118;--surface2:#16161f;--border:#252535;--border2:#2e2e45;"
        "--gold:#cfa347;--gold2:#f2d78f;--green:#3ecf6e;"
        "--text:#e7e4de;--muted:#948f89;--dim:#59544f"
        "}"
        "body{background:var(--bg);color:var(--text);"
        "font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;"
        "font-size:13px;min-height:100vh}"
        ".sf-header{background:var(--surface);border-bottom:2px solid var(--border2);"
        "padding:1rem 1.5rem;display:flex;align-items:center;gap:1.5rem;flex-wrap:wrap}"
        ".logo-area{display:flex;align-items:center;gap:14px}"
        ".logo-img{width:54px;height:54px;border-radius:13px;object-fit:cover;flex-shrink:0}"
        ".logo-name{font-size:26px;font-weight:900;letter-spacing:-.02em;"
        "background:linear-gradient(90deg,#8f6a28,#f2d78f);"
        "-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}"
        ".logo-sub{font-size:11px;color:var(--muted);margin-top:2px;letter-spacing:.03em}"
        ".sf-meta{display:flex;gap:.6rem;margin-left:auto;flex-wrap:wrap}"
        ".meta-chip{font-size:11.5px;color:var(--muted);display:flex;align-items:center;gap:5px;"
        "background:var(--surface2);border:1px solid var(--border);padding:5px 11px;border-radius:6px}"
        ".sf-stats{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;padding:1rem 1.5rem}"
        ".stat{background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:12px 16px}"
        ".stat-label{font-size:11px;color:var(--muted);margin-bottom:4px;text-transform:uppercase;letter-spacing:.05em}"
        ".stat-val{font-size:28px;font-weight:900;color:var(--text)}"
        ".stat-val.green{color:var(--green)}"
        ".sf-table-wrap{padding:0 1.5rem;overflow-x:auto;-webkit-overflow-scrolling:touch}"
        "table{width:100%;border-collapse:collapse;min-width:1100px;table-layout:fixed}"
        "col.c-num{width:46px}col.c-date{width:64px}col.c-tag{width:46px}"
        "col.c-short{width:108px}col.c-musik{width:100px}col.c-hook{width:158px}"
        "col.c-text{width:235px}col.c-titel{width:148px}col.c-yt{width:158px}"
        "col.c-ig{width:155px}col.c-status{width:104px}"
        "thead tr{background:var(--surface2)}"
        "thead th{padding:10px;text-align:left;font-size:13px;font-weight:700;"
        "letter-spacing:.04em;border:1px solid var(--border2);white-space:nowrap}"
        "thead th.center{text-align:center}"
        ".th-num,.th-date,.th-tag,.th-short,.th-musik,"
        ".th-hook,.th-text,.th-titel,.th-yt,.th-ig,.th-status{color:var(--muted)}"
        "tbody tr{border-bottom:1px solid var(--border);transition:background .15s,opacity .35s}"
        "tbody tr:hover{background:#13131e}"
        "tbody tr.done{opacity:.32}"
        "td{padding:13px 10px;vertical-align:top;border-left:1px solid var(--border);"
        "border-right:1px solid var(--border);word-wrap:break-word;overflow-wrap:break-word}"
        ".td-num{text-align:center;font-size:13px;font-weight:700;color:var(--text);vertical-align:middle}"
        ".td-date{text-align:center;font-size:13px;font-weight:700;color:var(--text);"
        "vertical-align:middle;white-space:nowrap}"
        ".td-tag{text-align:center;font-size:13px;font-weight:700;color:var(--text);vertical-align:middle}"
        ".td-short{text-align:center;font-size:12px;font-weight:400;color:var(--muted);"
        "vertical-align:middle;line-height:1.35}"
        ".td-musik{text-align:center;vertical-align:middle}"
        ".musik-badge{display:inline-block;padding:3px 9px;border-radius:5px;"
        "font-size:11px;font-weight:700;letter-spacing:.03em}"
        ".musik-phonk{background:#1a1028;color:#c084fc;border:1px solid #4a2070}"
        ".musik-action{background:#1c1010;color:#f87171;border:1px solid #6b2020}"
        ".musik-wissen{background:#0d1f14;color:#4ade80;border:1px solid #1a5228}"
        ".musik-clever{background:#1a180d;color:#fbbf24;border:1px solid #6b5020}"
        ".musik-track{display:block;font-size:10.5px;color:var(--muted);margin-top:4px;line-height:1.3}"
        ".hook-text{font-size:13.5px;font-weight:600;color:var(--text);line-height:1.5}"
        ".text-body{font-size:13px;font-weight:400;color:#ccc;line-height:1.6}"
        ".titel-text{font-size:13.5px;font-weight:600;color:var(--text);line-height:1.45}"
        ".desc-text{font-size:12px;font-weight:400;color:#bbb;line-height:1.55}"
        ".copy-btn{margin-top:6px;display:inline-flex;align-items:center;gap:4px;"
        "background:var(--surface2);border:1px solid var(--border2);color:var(--muted);"
        "font-size:10px;padding:3px 7px;border-radius:4px;cursor:pointer;transition:all .15s;"
        "white-space:nowrap;font-family:inherit}"
        ".copy-btn:hover{border-color:var(--gold);color:var(--gold2)}"
        ".copy-btn.copied{background:#0d2a1a;border-color:var(--green);color:var(--green)}"
        ".td-status{text-align:center;vertical-align:middle}"
        ".status-btn{display:inline-flex;align-items:center;gap:5px;padding:5px 10px;"
        "border-radius:6px;font-size:12px;font-weight:700;cursor:pointer;border:none;"
        "transition:all .2s;font-family:inherit}"
        ".status-btn.pending{background:var(--surface2);border:1px solid var(--border2);color:var(--dim)}"
        ".status-btn.pending:hover{border-color:var(--gold);color:var(--gold2)}"
        ".status-btn.fertig{background:#0d2a1a;border:1px solid #1a5228;color:var(--green)}"
        ".sf-footer{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;"
        "padding:.6rem 1.5rem .8rem}"
        ".footer-box{background:var(--surface);border:1px solid var(--border);"
        "border-radius:10px;padding:9px 14px}"
        ".footer-title{font-size:11px;font-weight:700;color:var(--gold);margin-bottom:7px;"
        "text-transform:uppercase;letter-spacing:.05em}"
        ".footer-list{list-style:none}"
        ".footer-list li{font-size:12px;color:var(--muted);padding:2px 0}"
        ".export-btns{display:flex;gap:6px;flex-wrap:wrap;margin-top:7px}"
        ".exp-btn{display:inline-flex;align-items:center;gap:5px;padding:4px 9px;"
        "border-radius:6px;font-size:11px;font-weight:600;cursor:pointer;border:none;font-family:inherit}"
        ".exp-btn.xls,.exp-btn.csv,.exp-btn.txt{background:var(--surface2);color:var(--muted);"
        "border:1px solid var(--border2)}"
        ".exp-btn.xls:hover,.exp-btn.csv:hover,.exp-btn.txt:hover{border-color:var(--gold);color:var(--gold2)}"
        ".hinweis-text{font-size:11.5px;color:var(--muted);line-height:1.5;margin-top:4px}"
        ".footer-brand{display:flex;flex-direction:row;align-items:center;gap:12px}"
        ".footer-brand-name{font-size:18px;font-weight:900;letter-spacing:-.02em;"
        "background:linear-gradient(90deg,#8f6a28,#f2d78f);"
        "-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}"
        ".footer-brand-slogan{font-size:11px;color:var(--muted);line-height:1.4;margin-top:2px}"
        ".field-copied{opacity:.42;transition:opacity .3s}"
        "@media(max-width:680px){"
        ".sf-footer{grid-template-columns:1fr 1fr}"
        ".copy-btn{padding:6px 10px;font-size:11px;min-height:34px}"
        ".status-btn{padding:8px 12px;min-height:38px}"
        ".sf-stats{gap:6px;padding:.8rem 1rem}"
        "}"
    )

    _BODY = (
        f'<div class="sf-header">'
        f'<div class="logo-area"><img class="logo-img" src="data:image/png;base64,{_LOGO_B64_54}" alt="ShortFlow"><div>'
        f'<div class="logo-name">ShortFlow</div>'
        f'<div class="logo-sub">Ideen. Inhalte. Shorts. Fertig.</div>'
        f'</div></div>'
        f'<div class="sf-meta">'
        f'<div class="meta-chip">&#128197; {topic_disp}</div>'
        f'<div class="meta-chip">&#9889; {count} Shorts</div>'
        f'<div class="meta-chip">&#128640; ShortFlow v1.0</div>'
        f'</div></div>'
        f'<div class="sf-stats">'
        f'<div class="stat"><div class="stat-label">Shorts gesamt</div>'
        f'<div class="stat-val" id="total-count">{count}</div></div>'
        f'<div class="stat"><div class="stat-label">Fertig</div>'
        f'<div class="stat-val green" id="done-count">0</div></div>'
        f'<div class="stat"><div class="stat-label">Ausstehend</div>'
        f'<div class="stat-val" id="pending-count">{count}</div></div>'
        f'</div>'
        f'<div class="sf-table-wrap"><table>'
        f'<colgroup>'
        f'<col class="c-num"><col class="c-date"><col class="c-tag"><col class="c-short"><col class="c-musik">'
        f'<col class="c-hook"><col class="c-text"><col class="c-titel"><col class="c-yt"><col class="c-ig">'
        f'<col class="c-status">'
        f'</colgroup>'
        f'<thead><tr>'
        f'<th class="th-num center">#</th>'
        f'<th class="th-date center">Datum</th>'
        f'<th class="th-tag center">Tag</th>'
        f'<th class="th-short center">Short</th>'
        f'<th class="th-musik center">Musik</th>'
        f'<th class="th-hook">&#9889; Hook</th>'
        f'<th class="th-text">&#127908; Textblock</th>'
        f'<th class="th-titel">&#128250; Titel</th>'
        f'<th class="th-yt">YT Beschreibung</th>'
        f'<th class="th-ig">IG Beschreibung</th>'
        f'<th class="th-status center">&#10003; Status</th>'
        f'</tr></thead>'
        f'<tbody id="table-body">{_tbody_html}</tbody>'
        f'</table></div>'
        f'<div class="sf-footer">'
        f'<div class="footer-box">'
        f'<div class="footer-title">&#8595; Export</div>'
        f'<p style="font-size:11px;color:var(--muted)">Automatisch von ShortFlow erstellt.</p>'
        f'<div class="export-btns">'
        f'<button class="exp-btn xls" onclick="exportXLS()">&#128202; Excel</button>'
        f'<button class="exp-btn csv" onclick="exportCSV()">&#128196; CSV</button>'
        f'<button class="exp-btn txt" onclick="exportTXT()">&#128203; Texte</button>'
        f'</div></div>'
        f'<div class="footer-box">'
        f'<div class="footer-title">&#9654; Verwendung</div>'
        f'<ul class="footer-list">'
        f'<li>1. Texte f&#252;r ElevenLabs verwenden</li>'
        f'<li>2. Videos mit Prompts generieren</li>'
        f'<li>3. Shorts bearbeiten &amp; hochladen</li>'
        f'</ul></div>'
        f'<div class="footer-box">'
        f'<div class="footer-title">&#9733; Hinweis</div>'
        f'<p class="hinweis-text">Alle Inhalte wurden KI-generiert und sollten vor Ver&#246;ffentlichung '
        f'gepr&#252;ft werden.</p></div>'
        f'<div class="footer-box footer-brand"><img class="logo-img" style="width:48px;height:48px" src="data:image/png;base64,{_LOGO_B64_70}" alt="ShortFlow"><div>'
        f'<div class="footer-brand-name">ShortFlow</div>'
        f'<div class="footer-brand-slogan">Erstelle bessere Shorts.<br>Spare Zeit. Erreiche mehr.</div>'
        f'</div></div>'
        f'</div>'
    )

    _project_id = _hashlib.sha256(str(Path(path).resolve()).encode("utf-8")).hexdigest()[:16]
    _sf_pfx_js = _json.dumps(f"sf_{_project_id}_")
    _SCRIPT = (
        f'const rows={rows_js};\n'
        f'var _sfPfx={_sf_pfx_js};\n'
        'function lsGet(k){try{return localStorage.getItem(k);}catch(e){return null;}}\n'
        'function lsSet(k,v){try{localStorage.setItem(k,v);}catch(e){}}\n'
        'function lsRemove(k){try{localStorage.removeItem(k);}catch(e){}}\n'
        'function updateCounts(){'
        'const done=document.querySelectorAll("#table-body tr.done").length;'
        'const all=document.querySelectorAll("#table-body tr").length;'
        'document.getElementById("done-count").textContent=done;'
        'document.getElementById("pending-count").textContent=all-done;}\n'
        'function fallbackCopy(text,btn){'
        'var isIOS=/ipad|iphone/i.test(navigator.userAgent);'
        'var ta=document.createElement("textarea");ta.value=text;'
        'ta.style.position="fixed";ta.style.top="0";ta.style.left="0";'
        'ta.style.width="100%";ta.style.height="1px";'
        'ta.style.opacity="0.01";ta.style.fontSize="16px";'
        'ta.style.padding="0";ta.style.border="none";'
        'document.body.appendChild(ta);'
        'if(isIOS){'
        'var r=document.createRange();r.selectNodeContents(ta);'
        'var s=window.getSelection();s.removeAllRanges();s.addRange(r);'
        'ta.setSelectionRange(0,999999);'
        '}else{ta.focus();ta.select();}'
        'try{document.execCommand("copy");}catch(e){}'
        'document.body.removeChild(ta);markCopied(btn);}\n'
        'function markCopied(btn){'
        'btn.classList.add("copied");btn.textContent="✓ Kopiert";'
        'const field=btn.previousElementSibling;'
        'if(field)field.classList.add("field-copied");}\n'
        'function doCopy(btn){'
        'var text=btn.getAttribute("data-copy");'
        'if(location.protocol==="https:"&&navigator.clipboard&&navigator.clipboard.writeText){'
        'navigator.clipboard.writeText(text).then(function(){markCopied(btn);}).catch(function(){fallbackCopy(text,btn);});'
        '}else{fallbackCopy(text,btn);}}\n'
        'function toggleStatus(btn){'
        'const tr=btn.closest("tr");const short=btn.dataset.short;'
        'if(tr.classList.contains("done")){'
        'tr.classList.remove("done");'
        'btn.className="status-btn pending";btn.textContent="Ausstehend";'
        'lsRemove(_sfPfx+short);'
        '}else{'
        'tr.classList.add("done");'
        'btn.className="status-btn fertig";btn.textContent="✓ Fertig";'
        'lsSet(_sfPfx+short,"1");}updateCounts();}\n'
        'function initTable(){'
        'document.querySelectorAll(".copy-btn[data-copy]").forEach(function(btn){'
        'btn.onclick=function(){doCopy(btn);};});'
        'document.querySelectorAll(".status-btn[data-short]").forEach(function(btn){'
        'const saved=lsGet(_sfPfx+btn.dataset.short);'
        'if(saved==="1"){const tr=btn.closest("tr");'
        'tr.classList.add("done");'
        'btn.className="status-btn fertig";btn.textContent="✓ Fertig";}'
        'btn.onclick=function(){toggleStatus(btn);};});'
        'updateCounts();}\n'
        'function exportXLS(){'
        'const cols=["#","Datum","Tag","Short","Musik","Hook","Textblock","Titel","YT Beschreibung","IG Beschreibung"];'
        'const keys=["num","date","tag","short","musik","hook","text","titel","yt","ig"];'
        'let xml=\'<?xml version="1.0"?>\\n<?mso-application progid="Excel.Sheet"?>\\n\';'
        'xml+=\'<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet" \';'
        'xml+=\'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">\\n\';'
        'xml+=\'<Worksheet ss:Name="Shorts">\\n<Table>\\n\';'
        'xml+="<Row>";cols.forEach(c=>{xml+=\'<Cell><Data ss:Type="String">\'+c.replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;")+\'</Data></Cell>\';});xml+="</Row>\\n";'
        'rows.forEach(r=>{xml+="<Row>";keys.forEach(k=>{const v=String(r[k]||"").replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");xml+=\'<Cell><Data ss:Type="String">\'+v+\'</Data></Cell>\';});xml+="</Row>\\n";});'
        'xml+=\'</Table>\\n</Worksheet>\\n</Workbook>\';'
        'const b=new Blob([xml],{type:"application/vnd.ms-excel"});'
        'const a=document.createElement("a");a.href=URL.createObjectURL(b);'
        f'a.download="ShortFlow_{topic_disp.replace(" ","_")}.xls";a.click();}}\n'
        f'function exportCSV(){{'
        'const h=["#","Datum","Tag","Short","Musik","Hook","Textblock","Titel","YT","IG"];'
        'const csv=[h.join(","),...rows.map(r=>'
        '[r.num,r.date,r.tag,r.short,r.musik,r.hook,r.text,r.titel,r.yt,r.ig]'
        '.map(v=>\'"\'+String(v).replace(/"/g,\'""\')+\'"\').join(","))];'
        'const b=new Blob(["\\uFEFF"+csv.join("\\n")],{type:"text/csv;charset=utf-8;"});'
        'const a=document.createElement("a");a.href=URL.createObjectURL(b);'
        f'a.download="ShortFlow_{topic_disp.replace(" ","_")}.csv";a.click();}}\n'
        f'function exportTXT(){{'
        'const t=rows.map(r=>'
        '`[${r.num}] ${r.date} ${r.tag} – ${r.short}\\nHOOK: ${r.hook}\\nTEXT: ${r.text}'
        '\\nTITEL: ${r.titel}\\nYT: ${r.yt}\\nIG: ${r.ig}`).join("\\n\\n---\\n\\n");'
        'const b=new Blob([t],{type:"text/plain;charset=utf-8"});'
        'const a=document.createElement("a");a.href=URL.createObjectURL(b);'
        f'a.download="ShortFlow_Texte_{topic_disp.replace(" ","_")}.txt";a.click();}}\n'
        'initTable();'
    )

    page = (
        '<!DOCTYPE html><html lang="de"><head>'
        '<meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        f'<title>ShortFlow — {topic_disp}</title>'
        f'<style>{_CSS}</style>'
        f'</head><body>{_BODY}'
        f'<script>{_SCRIPT}</script>'
        '</body></html>'
    )
    Path(path).write_text(page, encoding="utf-8")


def append_to_html(df: pd.DataFrame, path: str | Path, topic: str = "", music_config: dict | None = None) -> None:
    """If path exists, merge existing rows with new df and rewrite. Otherwise create fresh."""
    path = Path(path)
    if not path.exists():
        save_html(df, path, topic=topic, music_config=music_config)
        return
    try:
        existing_df = load_html(path)
    except Exception as exc:
        print(f"[MERGE] Bestehende HTML nicht lesbar ({exc}) — überschreibe mit neuer Tabelle")
        save_html(df, path, topic=topic, music_config=music_config)
        return
    merged = pd.concat([existing_df, df], ignore_index=True)
    save_html(merged, path, topic=topic, music_config=music_config)
    print(f"[MERGE] {len(existing_df)} bestehende + {len(df)} neue Rows → {len(merged)} gesamt")


def save_ig_html(df: pd.DataFrame, path: str | Path, topic: str = "") -> None:
    import html as _html_lib
    from datetime import date as _date

    topic_disp = topic.strip() or "ShortFlow"
    today = _date.today().strftime("%d.%m.%Y")
    count = len(df)

    def _he(s: str) -> str:
        return _html_lib.escape(str(s), quote=False)

    cards_html = []
    for i, (_, row) in enumerate(df.iterrows()):
        short = str(row.get("Short", ""))
        num = short.replace("Short", "")
        date_val = str(row.get("Datum", ""))
        titel = str(row.get("Titel", ""))
        ig = str(row.get("IGBeschreibung", ""))
        hook = str(row.get("Hook", ""))
        tags = _auto_tags(hook, titel, topic)
        ig_text = titel + '\n' + ig + '\n' + ' '.join(tags)
        cards_html.append(
            f'<div class="card">'
            f'<div class="card-meta">'
            f'<span class="num">Short {_html_lib.escape(num)}</span>'
            f'<span class="cdate">{_html_lib.escape(date_val)}</span>'
            f'</div>'
            f'<div class="titel">{_html_lib.escape(titel)}</div>'
            f'<button class="copy-btn" onclick="doCopy(this,{i})">Kopieren</button>'
            f'<textarea class="ig-ta" id="ta{i}" readonly>{_html_lib.escape(ig_text)}</textarea>'
            f'</div>'
        )

    cards = '\n'.join(cards_html)

    _CSS = (
        "*{box-sizing:border-box;margin:0;padding:0}"
        "body{background:#0b0b10;color:#e7e4de;"
        "font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Helvetica,Arial,sans-serif;"
        "padding:12px;max-width:620px;margin:0 auto}"
        "header{padding:14px 0 12px;border-bottom:1px solid #252535;margin-bottom:14px}"
        "h1{font-size:21px;font-weight:900;letter-spacing:-.01em;"
        "background:linear-gradient(90deg,#8f6a28,#f2d78f);"
        "-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}"
        ".hdr-meta{font-size:12px;color:#777;margin-top:5px}"
        ".ios-hint{background:#1a1a10;border:1px solid #3a3a10;border-radius:8px;"
        "padding:10px 12px;margin-bottom:14px;font-size:12px;color:#aaa;line-height:1.5}"
        ".ios-hint strong{color:#cfa347}"
        ".card{background:#111118;border:1px solid #252535;border-radius:12px;padding:14px;margin-bottom:13px}"
        ".card-meta{display:flex;align-items:center;gap:8px;margin-bottom:8px;flex-wrap:wrap}"
        ".num{background:#17130d;color:#cfa347;font-size:11px;font-weight:700;"
        "padding:3px 9px;border-radius:5px;white-space:nowrap}"
        ".cdate{color:#666;font-size:12px;white-space:nowrap}"
        ".titel{font-size:14px;font-weight:600;color:#e2e2e2;margin-bottom:10px;line-height:1.4}"
        ".copy-btn{display:block;width:100%;padding:14px;background:#16161f;"
        "border:1px solid #252535;color:#888;font-size:15px;font-weight:600;"
        "border-radius:8px;cursor:pointer;font-family:inherit;margin-bottom:10px;"
        "transition:background .15s,border-color .15s,color .15s;"
        "-webkit-tap-highlight-color:rgba(0,0,0,0)}"
        ".copy-btn.sel{background:#1a1a0d;border-color:#cfa347;color:#cfa347}"
        ".copy-btn.ok{background:#0d2a1a;border-color:#2ecc71;color:#2ecc71}"
        ".ig-ta{display:block;width:100%;min-height:140px;background:#0d0d14;"
        "border:1px solid #1e1e2e;color:#ccc;font-size:14px;line-height:1.7;"
        "padding:12px;border-radius:8px;font-family:inherit;resize:vertical;"
        "-webkit-user-select:text !important;user-select:text !important;cursor:text}"
        ".card.done{opacity:.5;transition:opacity .3s}"
        ".hint-error{display:none}"
        ".ios-hint.blocked{background:#1a0d0d;border-color:#c0392b}"
        ".ios-hint.blocked .hint-normal{display:none}"
        ".ios-hint.blocked .hint-error{display:block}"
    )

    _SCRIPT = (
        'function selAll(ta){'
        'ta.focus();ta.select();'
        'try{ta.setSelectionRange(0,ta.value.length);}catch(e){}}\n'
        'function onOk(btn){'
        'btn.classList.remove("sel","ok");btn.classList.add("ok");btn.textContent="✓ Kopiert";'
        'var c=btn.closest(".card");if(c)c.classList.add("done");}\n'
        'function showSafariHint(){'
        'var h=document.getElementById("ios-hint");'
        'if(h)h.classList.add("blocked");}\n'
        'function onFail(btn,ta){'
        'btn.classList.remove("sel","ok");btn.classList.add("sel");'
        'btn.textContent="Manuell kopieren";'
        'selAll(ta);showSafariHint();}\n'
        'function tryExec(btn,ta){'
        'var ok=false;try{ok=document.execCommand("copy");}catch(e){}'
        'if(ok){onOk(btn);}else{onFail(btn,ta);}}\n'
        'function doCopy(btn,idx){'
        'var ta=document.getElementById("ta"+idx);'
        'if(!ta)return;'
        'var text=ta.value;'
        'if(navigator.clipboard&&navigator.clipboard.writeText){'
        'navigator.clipboard.writeText(text)'
        '.then(function(){onOk(btn);})'
        '.catch(function(){selAll(ta);tryExec(btn,ta);});'
        '}else{selAll(ta);tryExec(btn,ta);}}\n'
        'document.querySelectorAll(".ig-ta").forEach(function(ta){'
        'ta.addEventListener("click",function(){selAll(ta);});});'
    )

    page = (
        '<!DOCTYPE html><html lang="de"><head>'
        '<meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1">'
        f'<title>IG Export — {_html_lib.escape(topic_disp)}</title>'
        f'<style>{_CSS}</style>'
        '</head><body>'
        '<header>'
        '<h1>IG Export</h1>'
        f'<div class="hdr-meta">{_html_lib.escape(topic_disp)} &nbsp;&middot;&nbsp; {_html_lib.escape(today)} &nbsp;&middot;&nbsp; {count} Shorts</div>'
        '</header>'
        '<div class="ios-hint" id="ios-hint">'
        '<span class="hint-normal"><strong>Safari:</strong> F&uuml;r 1-Tap-Kopieren Datei in Safari &ouml;ffnen: Teilen &rarr; &bdquo;In Safari &ouml;ffnen&ldquo;</span>'
        '<span class="hint-error"><strong>&#9888; Clipboard blockiert</strong> &mdash; du bist in der Dateien-Vorschau. Tippe Teilen &rarr; &bdquo;In Safari &ouml;ffnen&ldquo; und dann nochmal Kopieren.</span>'
        '</div>'
        f'{cards}'
        f'<script>{_SCRIPT}</script>'
        '</body></html>'
    )
    Path(path).write_text(page, encoding="utf-8")


def load_csv(path: str | Path) -> pd.DataFrame:
    df = pd.read_csv(path, sep=";", encoding="utf-8-sig")
    if "Short" not in df.columns:
        raise ValueError("CSV-Datei fehlt die Spalte 'Short'.")
    if "Status" not in df.columns:
        df["Status"] = "Ausstehend"
    return df


def load_xlsx(path: str | Path) -> pd.DataFrame:
    raw = pd.read_excel(path, dtype=str).fillna("")
    raw.columns = [c.strip().lower() for c in raw.columns]
    rows = []
    for i, (_, row) in enumerate(raw.iterrows(), start=1):
        fertig = str(row.get("fertig", "")).strip()
        # Support both old (beschreibung) and new (ytbeschreibung/igbeschreibung) column names
        yt = str(row.get("ytbeschreibung", row.get("beschreibung", ""))).strip()
        ig = str(row.get("igbeschreibung", "")).strip()
        rows.append({
            "Short":          str(row.get("short", f"Short{i:02d}")).strip() or f"Short{i:02d}",
            "Datum":          str(row.get("datum", "")).strip(),
            "Tag":            str(row.get("tag", "")).strip(),
            "Hook":           str(row.get("hook", "")),
            "Text":           str(row.get("text", "")),
            "Titel":          str(row.get("titel", "")),
            "YTBeschreibung": yt,
            "IGBeschreibung": ig,
            "Prompts":        str(row.get("prompts", "")),
            "Status":         "Fertig" if fertig == "Fertig" else "Ausstehend",
            "EnergiTyp":      str(row.get("energityp", "wissen")).strip().lower() or "wissen",
        })
    if not rows:
        raise ValueError("Keine Zeilen in der Excel-Datei gefunden.")
    from modules.brain import COLUMNS
    return pd.DataFrame(rows, columns=COLUMNS)


def save_prompts_json(df: pd.DataFrame, path: str | Path) -> None:
    import json
    path = Path(path)
    data = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
    data.update({str(row["Short"]): str(row.get("Prompts", "")) for _, row in df.iterrows()})
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_prompts_json(path: str | Path) -> dict:
    import json
    return json.loads(Path(path).read_text(encoding="utf-8"))


def load_html(path: str | Path) -> pd.DataFrame:
    import html as _html
    import json as _json
    import re as _re
    text = Path(path).read_text(encoding="utf-8")

    # New JS-based format (const rows=[...])
    m = _re.search(r'const rows=(\[.*?\]);', text, _re.DOTALL)
    if m:
        try:
            rows_data = _json.loads(m.group(1))
        except _json.JSONDecodeError as exc:
            raise ValueError(f"HTML-Datei: rows-JSON ungültig: {exc}") from None
        rows = []
        for i, r in enumerate(rows_data, 1):
            rows.append({
                "Short":          r.get("short", f"Short{i:02d}"),
                "Datum":          r.get("date", ""),
                "Tag":            r.get("tag", ""),
                "Hook":           r.get("hook", ""),
                "Text":           r.get("text", ""),
                "Titel":          r.get("titel", ""),
                "YTBeschreibung": r.get("yt", ""),
                "IGBeschreibung": r.get("ig", ""),
                "Prompts":        "",
                "Status":         "Ausstehend",
                "EnergiTyp":      r.get("musik", "wissen"),
            })
        if not rows:
            raise ValueError("Keine Shorts in der HTML-Datei gefunden.")
        from modules.brain import COLUMNS
        return pd.DataFrame(rows, columns=COLUMNS)

    # Legacy format (old save_html with data-row attributes)
    _COLS_LEGACY = ["Short", "Hook", "Text", "Titel", "YTBeschreibung", "IGBeschreibung",
                    "Datum", "Tag", "Prompts", "Status", "EnergiTyp"]
    rows = []
    for m2 in _re.finditer(r'<tr[^>]+data-row="(\d+)"[^>]*>(.*?)</tr>', text, _re.DOTALL):
        idx = int(m2.group(1))
        r = m2.group(2)
        short_m = _re.search(r'<td class="cs"[^>]*>([^<]*)</td>', r)
        short = short_m.group(1).strip() if short_m else f"Short{idx+1:02d}"
        hook_m = _re.search(r'data-val="([^"]*)"', r)
        hook = _html.unescape(hook_m.group(1)) if hook_m else ""
        bts = _re.findall(r'class="bt"[^>]+data-val="([^"]*)"', r)
        text_val = _html.unescape(bts[0]) if len(bts) > 0 else ""
        titel    = _html.unescape(bts[1]) if len(bts) > 1 else ""
        besch    = _html.unescape(bts[2]) if len(bts) > 2 else ""
        rows.append({
            "Short": short, "Hook": hook, "Text": text_val,
            "Titel": titel, "YTBeschreibung": besch, "IGBeschreibung": "",
            "Datum": "", "Tag": "", "Prompts": "", "Status": "Ausstehend", "EnergiTyp": "wissen",
        })
    if not rows:
        raise ValueError("Keine Shorts in der HTML-Datei gefunden.")
    from modules.brain import COLUMNS
    return pd.DataFrame(rows, columns=COLUMNS)


# ── Internal ──────────────────────────────────────────────────────────────────

def _prepare(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "Status" in df.columns and "Fertig" not in df.columns:
        df["Fertig"] = df["Status"].apply(
            lambda v: "Fertig" if str(v).strip().lower() == "fertig" else "Offen"
        )
    for col in _DATA_COLS:
        if col not in df.columns:
            df[col] = ""
    return df[_DATA_COLS]


def _border(color: str = _BORDER, weight: str = "thin") -> Border:
    s = Side(style=weight, color=color)
    return Border(right=s, bottom=s)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _write_header(ws) -> None:
    accent_bottom = Side(style="medium", color=_TAB_COLOR)
    right         = Side(style="thin", color=_BORDER)

    for ci, col in enumerate(_DISPLAY_COLS, start=1):
        cell = ws.cell(row=1, column=ci, value=col.upper())
        cell.font      = Font(name="Calibri", bold=True, size=10,
                              color=_COL_ACCENT.get(col, "FFFFFF"))
        cell.fill      = _fill(_BG_HEADER)
        cell.alignment = Alignment(
            horizontal="center" if col == "#" else "left",
            vertical="center",
        )
        cell.border = Border(bottom=accent_bottom, right=right)

    ws.row_dimensions[1].height = 38


def _write_rows(ws, df: pd.DataFrame) -> None:
    bdr = _border()

    for ri, (_, row) in enumerate(df.iterrows(), start=2):
        bg       = _BG_ROW_A if ri % 2 == 0 else _BG_ROW_B
        row_fill = _fill(bg)

        # ── # column ─────────────────────────────────────────────────────────
        num = ws.cell(row=ri, column=1, value=f"{ri - 1:02d}")
        num.font      = Font(name="Calibri", bold=True, size=13, color=_TEXT_NUM)
        num.fill      = row_fill
        num.alignment = Alignment(horizontal="center", vertical="center")
        num.border    = bdr

        # ── data columns ─────────────────────────────────────────────────────
        for ci, col in enumerate(_DATA_COLS, start=2):
            raw   = row.get(col, "")
            value = str(raw) if pd.notna(raw) else ""
            cell  = ws.cell(row=ri, column=ci, value=value)
            cell.border = bdr

            if col == "Fertig":
                done = value.strip() == "Fertig"
                cell.font      = Font(name="Calibri", bold=done, size=10,
                                      color=_FG_FERTIG if done else _FG_OFFEN)
                cell.fill      = _fill(_BG_FERTIG) if done else row_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif col == "Hook":
                cell.font      = Font(name="Calibri", bold=True, size=12,
                                      color=_TEXT_HOOK)
                cell.fill      = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True,
                                           indent=1)

            else:
                cell.font      = Font(name="Calibri", size=10, color=_TEXT_BODY)
                cell.fill      = row_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True,
                                           indent=1)

        ws.row_dimensions[ri].height = 110


def _apply_col_widths(ws) -> None:
    for ci, col in enumerate(_DISPLAY_COLS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTHS.get(col, 20)


def _add_status_dropdown(ws, row_count: int) -> None:
    col_letter = get_column_letter(len(_DISPLAY_COLS))
    dv = DataValidation(
        type="list",
        formula1='"Offen,Fertig"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.sqref = f"{col_letter}2:{col_letter}{row_count + 1}"
    ws.add_data_validation(dv)


def _add_status_conditional_format(ws, row_count: int) -> None:
    col_letter = get_column_letter(len(_DISPLAY_COLS))
    cell_range = f"{col_letter}2:{col_letter}{row_count + 1}"
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(
            operator="equal",
            formula=['"Fertig"'],
            fill=_fill(_BG_FERTIG),
            font=Font(color=_FG_FERTIG, bold=True),
        ),
    )


def validate_html(path: str | Path) -> list[str]:
    errors: list[str] = []
    p = Path(path)
    if not p.exists():
        return ["Datei existiert nicht"]
    if p.stat().st_size == 0:
        errors.append("Datei ist leer (0 Bytes)")
    try:
        content = p.read_text(encoding="utf-8")
    except Exception as exc:
        return [f"Datei nicht lesbar: {exc}"]
    if "<tbody" not in content:
        errors.append("Kein <tbody> gefunden")
    if "<tr" not in content:
        errors.append("Keine <tr>-Zeilen gefunden")
    if "Short" not in content:
        errors.append("Keine Short-Daten gefunden")
    if "fetch(" in content:
        errors.append("Ungültig: enthält fetch()")
    return errors
